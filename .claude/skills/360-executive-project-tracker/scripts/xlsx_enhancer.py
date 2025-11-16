"""
XLSX Schema Enhancer
Adds Category and Source Details columns to make XLSX compatible with HTML dashboard
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def enhance_xlsx_schema(xlsx_path, output_path=None):
    """
    Add Category (Column O) and Source Details (Column P) to XLSX
    Maintains all existing data and formatting
    """
    if output_path is None:
        output_path = xlsx_path

    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Project Status']

    # Check if columns already exist
    headers = [cell.value for cell in ws[1]]

    # Add Category column (O) if not exists
    if len(headers) < 15 or headers[14] != 'Category':
        print("Adding Category column (O)...")
        ws.cell(1, 15, 'Category')

        # Style header
        header_cell = ws.cell(1, 15)
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column width
        ws.column_dimensions['O'].width = 25

        # Add data validation
        dv_category = DataValidation(
            type="list",
            formula1='"Partnership & Legal,Technical & Product,Pilot & Implementation,Business Development,Operations,Strategy & Planning"',
            allow_blank=True
        )
        dv_category.error = 'Please select a valid category'
        dv_category.errorTitle = 'Invalid Category'
        ws.add_data_validation(dv_category)
        dv_category.add(f'O2:O1000')

        # Auto-categorize existing tasks based on keywords
        category_keywords = {
            'Partnership & Legal': ['partnership', 'legal', 'contract', 'agreement', 'jv', 'joint venture'],
            'Technical & Product': ['api', 'integration', 'development', 'technical', 'testing', 'workflow'],
            'Pilot & Implementation': ['pilot', 'deployment', 'rollout', 'implementation', 'clinic'],
            'Business Development': ['proposal', 'pricing', 'sales', 'revenue', 'client', 'customer'],
            'Operations': ['operations', 'process', 'workflow', 'internal', 'team'],
            'Strategy & Planning': ['strategy', 'planning', 'roadmap', 'vision', 'assessment']
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=1000), start=2):
            description = row[2].value  # Column C (Task Description)
            if description:
                description_lower = description.lower()

                # Find best matching category
                for category, keywords in category_keywords.items():
                    if any(keyword in description_lower for keyword in keywords):
                        ws.cell(row_idx, 15, category)
                        break

        print("✓ Category column added with auto-categorization")
    else:
        print("Category column already exists")

    # Add Source Details column (P) if not exists
    if len(headers) < 16 or (len(headers) >= 16 and headers[15] != 'Source Details'):
        print("Adding Source Details column (P)...")
        ws.cell(1, 16, 'Source Details')

        # Style header
        header_cell = ws.cell(1, 16)
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column width
        ws.column_dimensions['P'].width = 60

        # Initialize with expanded version of Source Context
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=1000), start=2):
            source_context = row[9].value  # Column J (Source Context)
            if source_context and not row[15].value:  # Column P (Source Details)
                # Create detailed description from source context
                ws.cell(row_idx, 16, f"Details from: {source_context}")

        print("✓ Source Details column added")
    else:
        print("Source Details column already exists")

    # Save workbook
    print(f"Saving enhanced workbook to: {output_path}")
    wb.save(output_path)
    print("✓ Schema enhancement complete")

    return output_path

def verify_schema(xlsx_path):
    """
    Verify XLSX has all required columns for HTML dashboard
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Project Status']

    required_columns = {
        1: 'Task ID',
        2: 'Task Owner',
        3: 'Task Description',
        4: 'Status',
        5: 'Priority',
        6: 'Progress %',
        7: 'Blocker Details',
        8: 'Days Blocked',
        9: 'Last Update Date',
        10: 'Source Context',
        11: 'Next Steps',
        12: 'First Detected',
        13: 'Last Auto-Update',
        14: 'Manual Override',
        15: 'Category',
        16: 'Source Details'
    }

    headers = [cell.value for cell in ws[1]]

    print("\nSchema Verification:")
    print("=" * 60)

    all_good = True
    for col_num, expected_name in required_columns.items():
        actual_name = headers[col_num - 1] if col_num <= len(headers) else None
        status = "✓" if actual_name == expected_name else "✗"

        if actual_name != expected_name:
            all_good = False

        print(f"{status} Column {col_num:2d}: {expected_name:20s} -> {actual_name or 'MISSING'}")

    print("=" * 60)

    if all_good:
        print("✓ All required columns present")
    else:
        print("✗ Schema incomplete - run enhance_xlsx_schema() to fix")

    return all_good

if __name__ == "__main__":
    import sys

    # Check if path provided
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = '/home/claude/360_project_tracker.xlsx'

    # Verify current schema
    print("Checking current schema...")
    verify_schema(xlsx_path)

    # Enhance schema
    print("\nEnhancing schema...")
    output_path = xlsx_path.replace('.xlsx', '_enhanced.xlsx')
    enhance_xlsx_schema(xlsx_path, output_path)

    # Verify enhanced schema
    print("\nVerifying enhanced schema...")
    verify_schema(output_path)

    print(f"\nEnhanced file saved to: {output_path}")
