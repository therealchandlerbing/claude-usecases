"""
360 Project Status Tracker - Main Orchestrator
Integrates tracker creation, data collection, and dual-mode operation
"""

import sys
import os
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.comments import Comment

# Import our modules
from project_tracker_builder import ProjectTrackerBuilder
from data_collector import DataCollector, TaskExtractor

class TrackerOrchestrator:
    """Main orchestrator for the 360 Project Status Tracker system"""

    def __init__(self, mode='on_demand', config_path='/home/claude/config.json'):
        """
        Initialize orchestrator.

        Args:
            mode: 'on_demand' or 'recurring'
            config_path: Path to config.json file
        """
        self.mode = mode
        self.builder = ProjectTrackerBuilder()
        self.tracker_path = None

        # Load configuration from file
        try:
            with open(config_path, 'r') as f:
                self.full_config = json.load(f)

            # Extract commonly used config
            self.config = {
                'days_back': self.full_config['data_collection']['days_back'],
                'update_existing': self.full_config['tracker_settings']['update_existing'],
                'preserve_manual_edits': self.full_config['tracker_settings']['preserve_manual_edits'],
                'team_members': self.full_config['team']['members'],
                'drive_folders': self.full_config['data_collection']['sources']['drive']['folders']
            }
        except FileNotFoundError:
            # Fallback to default config if file not found
            self.full_config = None
            self.config = {
                'days_back': 7,
                'update_existing': True,
                'preserve_manual_edits': True,
                'team_members': [
                    'Chandler Lewis',
                    'Team Member',
                    'Unassigned'
                ],
                'drive_folders': [
                    {'name': 'A - Processed Meeting Deliverables'},
                    {'name': 'Fathom meeting deliverables', 'folder_id': '1P8gGXnNvCjvejFosFqTE7qjDVS3uOrmL'}
                ]
            }

    def run(self, days_back=7, create_new=False, use_sample_data=False):
        """
        Main execution flow for tracker generation/update.

        Args:
            days_back: How many days to look back for data
            create_new: Force creation of new tracker vs updating existing
            use_sample_data: Use sample data instead of live collection

        Returns:
            Path to generated/updated tracker
        """
        print("\n" + "=" * 70)
        print("360 PROJECT STATUS TRACKER")
        print(f"Mode: {self.mode.upper()}")
        print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 70)

        # Step 1: Create or load tracker
        if create_new or not self.tracker_path:
            print("\n[1/5] Creating tracker template...")
            self.tracker_path = self.builder.create_tracker_template()
            print(f"      ✓ Template created: {self.tracker_path}")

            if use_sample_data:
                print("      ✓ Adding sample data...")
                self.builder.add_sample_data(self.tracker_path)
        else:
            print(f"\n[1/5] Loading existing tracker: {self.tracker_path}")

        # Step 2: Collect data from sources
        print("\n[2/5] Collecting data from sources...")
        print(f"      • Looking back {days_back} days")

        if use_sample_data:
            print("      • Using sample data (live collection not executed)")
            tasks = []
        else:
            tasks = self.collect_live_data(days_back)
            print(f"      ✓ Collected {len(tasks)} tasks")

        # Step 3: Update tracker with collected data
        if tasks:
            print("\n[3/5] Updating tracker...")
            self.update_tracker_with_tasks(tasks)
            print(f"      ✓ {len(tasks)} tasks processed")
        else:
            print("\n[3/5] No new tasks to add (using existing data)")

        # Step 4: Add metadata
        print("\n[4/5] Updating metadata...")
        self.update_metadata()
        print("      ✓ Timestamps and mode updated")

        # Step 5: Save and finalize
        print("\n[5/5] Finalizing tracker...")
        print(f"      ✓ Saved to: {self.tracker_path}")

        # Generate summary
        summary = self.generate_summary()

        print("\n" + "=" * 70)
        print("TRACKER SUMMARY")
        print("=" * 70)
        for key, value in summary.items():
            print(f"  {key}: {value}")

        print("\n" + "=" * 70)
        print(f"✓ Tracker ready: {self.tracker_path}")
        print("=" * 70 + "\n")

        return self.tracker_path

    def collect_live_data(self, days_back=7):
        """
        Collect live data from Gmail, Calendar, Drive, and Asana.
        This is where we'll plug in the actual Claude tool calls.

        Returns:
            List of task dictionaries
        """
        # This function will be implemented with actual tool calls
        # For now, return structure showing what we'll collect

        drive_folders = self.config.get('drive_folders', [
            {'name': 'A - Processed Meeting Deliverables'},
            {'name': 'Fathom meeting deliverables', 'folder_id': '1P8gGXnNvCjvejFosFqTE7qjDVS3uOrmL'}
        ])

        asana_projects = self.full_config.get('data_collection', {}).get('sources', {}).get('asana', {}).get('projects', []) if self.full_config else []

        collection_summary = {
            'gmail_searches': [
                'status update after:YYYY/MM/DD',
                'blocker after:YYYY/MM/DD',
                'working on after:YYYY/MM/DD',
            ],
            'calendar_queries': {
                'time_min': (datetime.now() - timedelta(days=days_back)).isoformat() + 'Z',
                'time_max': datetime.now().isoformat() + 'Z',
            },
            'drive_search': {
                'folders': drive_folders,
                'date_filter': (datetime.now() - timedelta(days=days_back)).strftime('%Y-%m-%d')
            },
            'asana_search': {
                'projects': asana_projects,
                'modified_after': (datetime.now() - timedelta(days=days_back)).isoformat() + 'Z'
            }
        }

        print(f"      • Gmail: {len(collection_summary['gmail_searches'])} queries planned")
        print(f"      • Calendar: Last {days_back} days")
        print(f"      • Drive: {len(drive_folders)} folders:")
        for folder in drive_folders:
            folder_ref = f"'{folder['name']}'"
            if folder.get('folder_id'):
                folder_ref += f" (ID: {folder['folder_id']})"
            print(f"        - {folder_ref}")

        if asana_projects:
            print(f"      • Asana: {len(asana_projects)} projects:")
            for project in asana_projects:
                project_ref = f"'{project['name']}'"
                if project.get('project_id'):
                    project_ref += f" (ID: {project['project_id']})"
                print(f"        - {project_ref}")

        # Return empty for now (will be filled with actual collected tasks)
        return []

    def update_tracker_with_tasks(self, tasks):
        """
        Add or update tasks in the tracker.
        Respects manual override flags.
        """
        wb = load_workbook(self.tracker_path)
        ws = wb['Project Status']

        # Get existing tasks to check for updates
        existing_tasks = self.get_existing_tasks(ws)

        # Process each new task
        for task in tasks:
            task_id = task.get('task_id', '')

            if task_id in existing_tasks:
                # Check manual override
                existing_row = existing_tasks[task_id]['row']
                manual_override = ws[f'N{existing_row}'].value

                if manual_override == 'Yes':
                    print(f"      ⊘ Skipping {task_id} (manual override)")
                    continue
                else:
                    print(f"      ↻ Updating {task_id}")
                    self.update_task_row(ws, existing_row, task)
            else:
                print(f"      + Adding new task: {task_id}")
                self.add_new_task_row(ws, task)

        wb.save(self.tracker_path)

    def get_existing_tasks(self, worksheet):
        """Get dictionary of existing tasks with their row numbers"""
        existing = {}

        for row in range(2, worksheet.max_row + 1):
            task_id = worksheet[f'A{row}'].value
            if task_id:
                existing[task_id] = {
                    'row': row,
                    'data': [worksheet.cell(row=row, column=col).value
                            for col in range(1, 15)]
                }

        return existing

    def update_task_row(self, worksheet, row_num, task):
        """Update an existing task row with new data"""
        # Update only non-manual fields
        worksheet[f'C{row_num}'] = task.get('description', worksheet[f'C{row_num}'].value)
        worksheet[f'D{row_num}'] = task.get('status', worksheet[f'D{row_num}'].value)
        worksheet[f'G{row_num}'] = task.get('blocker_details', worksheet[f'G{row_num}'].value)
        worksheet[f'I{row_num}'] = datetime.now().strftime('%Y-%m-%d')
        worksheet[f'M{row_num}'] = datetime.now().strftime('%Y-%m-%d')

        # Add source context as comment
        source_cell = worksheet[f'J{row_num}']
        if task.get('source_context'):
            existing_comment = source_cell.comment.text if source_cell.comment else ''
            new_comment = f"{existing_comment}\n---\nUpdate {datetime.now().strftime('%Y-%m-%d')}: {task['source_context']}"
            source_cell.comment = Comment(new_comment, 'Tracker Auto-Update')

    def add_new_task_row(self, worksheet, task):
        """Add a new task row to the tracker"""
        next_row = worksheet.max_row + 1

        # Generate task ID if not provided
        if not task.get('task_id'):
            task['task_id'] = f"T{str(next_row - 1).zfill(3)}"

        # Write task data
        worksheet[f'A{next_row}'] = task.get('task_id', '')
        worksheet[f'B{next_row}'] = task.get('owner', 'Unassigned')
        worksheet[f'C{next_row}'] = task.get('description', '')
        worksheet[f'D{next_row}'] = task.get('status', 'Not Started')
        worksheet[f'E{next_row}'] = task.get('priority', 'Medium')
        worksheet[f'F{next_row}'] = task.get('progress', 0)
        worksheet[f'G{next_row}'] = task.get('blocker_details', '')
        worksheet[f'I{next_row}'] = datetime.now().strftime('%Y-%m-%d')
        worksheet[f'J{next_row}'] = task.get('source_context', '')
        worksheet[f'K{next_row}'] = task.get('next_steps', '')
        worksheet[f'L{next_row}'] = datetime.now().strftime('%Y-%m-%d')
        worksheet[f'M{next_row}'] = datetime.now().strftime('%Y-%m-%d')
        worksheet[f'N{next_row}'] = 'No'

        # Add source context as comment
        if task.get('source_context'):
            source_cell = worksheet[f'J{next_row}']
            source_cell.comment = Comment(task['source_context'], 'Tracker Auto-Update')

    def update_metadata(self):
        """Update tracker metadata (timestamps, mode, etc.)"""
        wb = load_workbook(self.tracker_path)
        dashboard = wb['Dashboard']

        # Update timestamp
        dashboard['B15'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Update mode
        dashboard['B16'] = self.mode.replace('_', ' ').title()

        wb.save(self.tracker_path)

    def generate_summary(self):
        """Generate summary statistics from current tracker"""
        wb = load_workbook(self.tracker_path, data_only=True)
        ws = wb['Project Status']

        summary = {
            'Total Tasks': 0,
            'Not Started': 0,
            'In Progress': 0,
            'Blocked': 0,
            'Done': 0,
            'Critical Priority': 0,
            'Blocked >3 Days': 0,
        }

        for row in range(2, ws.max_row + 1):
            task_id = ws[f'A{row}'].value
            if task_id:
                summary['Total Tasks'] += 1

                status = ws[f'D{row}'].value
                if status:
                    summary[status] = summary.get(status, 0) + 1

                priority = ws[f'E{row}'].value
                if priority == 'Critical':
                    summary['Critical Priority'] += 1

                days_blocked = ws[f'H{row}'].value
                if days_blocked and days_blocked > 3:
                    summary['Blocked >3 Days'] += 1

        wb.close()
        return summary

    def configure_recurring(self, schedule='weekly', day='Monday', time='09:00'):
        """
        Configure recurring mode settings.
        This would be used to set up automated runs.

        Args:
            schedule: 'weekly', 'biweekly', etc.
            day: Day of week for recurring run
            time: Time of day (HH:MM format)
        """
        self.mode = 'recurring'
        self.config['recurring'] = {
            'schedule': schedule,
            'day': day,
            'time': time,
            'enabled': True
        }

        print(f"\nRecurring mode configured:")
        print(f"  • Schedule: {schedule}")
        print(f"  • Day: {day}")
        print(f"  • Time: {time}")
        print(f"  • Next run: [Would calculate based on schedule]")


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='360 Project Status Tracker')
    parser.add_argument('--mode', choices=['on_demand', 'recurring'],
                       default='on_demand', help='Operating mode')
    parser.add_argument('--days', type=int, default=7,
                       help='Days to look back for data collection')
    parser.add_argument('--new', action='store_true',
                       help='Create new tracker instead of updating existing')
    parser.add_argument('--sample', action='store_true',
                       help='Use sample data instead of live collection')

    args = parser.parse_args()

    # Create orchestrator
    orchestrator = TrackerOrchestrator(mode=args.mode)

    # Run tracker generation/update
    tracker_path = orchestrator.run(
        days_back=args.days,
        create_new=args.new,
        use_sample_data=args.sample
    )

    # If recurring mode, show configuration options
    if args.mode == 'recurring':
        print("\nRecurring mode selected. Configuration options:")
        print("  • Weekly: Every Monday at 9am")
        print("  • Bi-weekly: Every other Monday")
        print("  • Custom: Set your own schedule")

if __name__ == "__main__":
    main()
