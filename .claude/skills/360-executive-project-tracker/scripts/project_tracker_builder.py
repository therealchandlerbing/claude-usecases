"""
360 Social Impact Studios - Project Status Tracker Builder
Dual-mode system: On-demand pulls and recurring updates
"""

import os
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import pandas as pd

class ProjectTrackerBuilder:
    def __init__(self):
        self.team_members = []
        self.tasks = []
        self.tracker_path = None
        self.mode = 'on_demand'  # or 'recurring'
        self.last_update = None

    def create_tracker_template(self, filepath='/mnt/user-data/outputs/360_project_tracker.xlsx'):
        """Create the base tracker template with all formatting and features"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Status"

        # Define colors
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        # Status colors
        blocked_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        in_progress_fill = PatternFill(start_color='FFE66D', end_color='FFE66D', fill_type='solid')
        done_fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
        not_started_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

        # Priority colors
        critical_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
        high_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')

        # Headers
        headers = [
            'Task ID', 'Task Owner', 'Task Description', 'Status', 'Priority',
            'Progress %', 'Blocker Details', 'Days Blocked', 'Last Update Date',
            'Source Context', 'Next Steps', 'First Detected', 'Last Auto-Update',
            'Manual Override'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        column_widths = {
            'A': 10,  # Task ID
            'B': 15,  # Task Owner
            'C': 40,  # Task Description
            'D': 12,  # Status
            'E': 10,  # Priority
            'F': 10,  # Progress %
            'G': 35,  # Blocker Details
            'H': 12,  # Days Blocked
            'I': 15,  # Last Update Date
            'J': 20,  # Source Context
            'K': 35,  # Next Steps
            'L': 15,  # First Detected
            'M': 15,  # Last Auto-Update
            'N': 15   # Manual Override
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze panes (freeze first row and first two columns)
        ws.freeze_panes = 'C2'

        # Add data validation for Status (column D)
        status_validation = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Blocked,Done"',
            allow_blank=False
        )
        status_validation.error = 'Invalid status'
        status_validation.errorTitle = 'Invalid Entry'
        ws.add_data_validation(status_validation)
        status_validation.add('D2:D1000')

        # Add data validation for Priority (column E)
        priority_validation = DataValidation(
            type="list",
            formula1='"Low,Medium,High,Critical"',
            allow_blank=False
        )
        priority_validation.error = 'Invalid priority'
        priority_validation.errorTitle = 'Invalid Entry'
        ws.add_data_validation(priority_validation)
        priority_validation.add('E2:E1000')

        # Add data validation for Manual Override (column N)
        override_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        ws.add_data_validation(override_validation)
        override_validation.add('N2:N1000')

        # Add conditional formatting for Status column
        ws.conditional_formatting.add('D2:D1000',
            CellIsRule(operator='equal', formula=['"Blocked"'], fill=blocked_fill))
        ws.conditional_formatting.add('D2:D1000',
            CellIsRule(operator='equal', formula=['"In Progress"'], fill=in_progress_fill))
        ws.conditional_formatting.add('D2:D1000',
            CellIsRule(operator='equal', formula=['"Done"'], fill=done_fill))
        ws.conditional_formatting.add('D2:D1000',
            CellIsRule(operator='equal', formula=['"Not Started"'], fill=not_started_fill))

        # Add conditional formatting for Priority column
        ws.conditional_formatting.add('E2:E1000',
            CellIsRule(operator='equal', formula=['"Critical"'], fill=critical_fill, font=Font(bold=True)))
        ws.conditional_formatting.add('E2:E1000',
            CellIsRule(operator='equal', formula=['"High"'], fill=high_fill))

        # Add conditional formatting for rows blocked >3 days (orange highlight)
        orange_fill = PatternFill(start_color='FFB84D', end_color='FFB84D', fill_type='solid')
        ws.conditional_formatting.add('A2:N1000',
            FormulaRule(formula=['AND($D2="Blocked",$H2>3,$H2<=7)'], fill=orange_fill))

        # Add conditional formatting for rows blocked >7 days (red highlight)
        red_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
        ws.conditional_formatting.add('A2:N1000',
            FormulaRule(formula=['AND($D2="Blocked",$H2>7)'], fill=red_fill, font=Font(bold=True, color='FFFFFF')))

        # Add formulas for calculated columns
        # Days Blocked formula (column H): IF(Status="Blocked", TODAY()-LastUpdate, "")
        for row in range(2, 1001):
            ws[f'H{row}'] = f'=IF(D{row}="Blocked",TODAY()-I{row},"")'

        # Create Summary Dashboard sheet
        summary = wb.create_sheet("Dashboard", 0)

        # Dashboard headers
        summary['A1'] = '360 Project Status Dashboard'
        summary['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        summary['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        summary.merge_cells('A1:D1')

        # Summary metrics
        summary['A3'] = 'Metric'
        summary['B3'] = 'Count'
        summary['A3'].font = Font(bold=True)
        summary['B3'].font = Font(bold=True)

        metrics = [
            ('Total Tasks', '=COUNTA(\'Project Status\'!A2:A1000)'),
            ('Not Started', '=COUNTIF(\'Project Status\'!D:D,"Not Started")'),
            ('In Progress', '=COUNTIF(\'Project Status\'!D:D,"In Progress")'),
            ('Blocked', '=COUNTIF(\'Project Status\'!D:D,"Blocked")'),
            ('Done', '=COUNTIF(\'Project Status\'!D:D,"Done")'),
            ('Critical Priority', '=COUNTIF(\'Project Status\'!E:E,"Critical")'),
            ('High Priority', '=COUNTIF(\'Project Status\'!E:E,"High")'),
            ('Blocked >3 Days', '=COUNTIFS(\'Project Status\'!D:D,"Blocked",\'Project Status\'!H:H,">3")'),
            ('Blocked >7 Days', '=COUNTIFS(\'Project Status\'!D:D,"Blocked",\'Project Status\'!H:H,">7")'),
        ]

        for idx, (metric, formula) in enumerate(metrics, start=4):
            summary[f'A{idx}'] = metric
            summary[f'B{idx}'] = formula

        # % Complete formula
        summary['A13'] = '% Complete'
        summary['B13'] = '=IFERROR(B8/(B4-B8),0)'
        summary['B13'].number_format = '0.0%'
        summary['A13'].font = Font(bold=True, size=12)
        summary['B13'].font = Font(bold=True, size=12)

        # Last update timestamp
        summary['A15'] = 'Last Updated'
        summary['B15'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        summary['A15'].font = Font(bold=True)

        # Mode indicator
        summary['A16'] = 'Update Mode'
        summary['B16'] = 'On-Demand'
        summary['A16'].font = Font(bold=True)

        # Column widths for dashboard
        summary.column_dimensions['A'].width = 20
        summary.column_dimensions['B'].width = 15

        # Add Instructions sheet
        instructions = wb.create_sheet("Instructions")

        instructions['A1'] = '360 Project Status Tracker - User Guide'
        instructions['A1'].font = Font(bold=True, size=14)

        guide_text = [
            '',
            'SYSTEM OVERVIEW',
            'This tracker operates in two modes:',
            '  • On-Demand: Manual pulls when you need status snapshot',
            '  • Recurring: Automatic weekly updates (coming soon)',
            '',
            'HOW TO USE',
            '1. Status column has dropdown: Not Started, In Progress, Blocked, Done',
            '2. Priority column has dropdown: Low, Medium, High, Critical',
            '3. Progress % shows visual data bars (enter 0-100)',
            '4. Manual Override column: Set to "Yes" to prevent auto-updates',
            '',
            'COLOR CODING',
            '  • Red: Blocked status or Critical priority',
            '  • Yellow: In Progress',
            '  • Green: Done',
            '  • Gray: Not Started',
            '  • Orange background: Blocked >3 days',
            '  • Red background: Blocked >7 days',
            '',
            'CELL COMMENTS',
            'Hover over Source Context cells to see details from:',
            '  • Gmail messages',
            '  • Calendar meetings',
            '  • Google Drive documents',
            '',
            'DASHBOARD',
            'See Dashboard tab for quick overview of:',
            '  • Total task counts by status',
            '  • Priority breakdowns',
            '  • Blocker alerts',
            '  • Completion percentage',
            '',
            'PROTECTING YOUR EDITS',
            'Set Manual Override to "Yes" for any task you edit manually.',
            'This prevents auto-updates from overwriting your changes.',
        ]

        for idx, line in enumerate(guide_text, start=1):
            instructions[f'A{idx+1}'] = line

        instructions.column_dimensions['A'].width = 70

        # Save the workbook
        wb.save(filepath)
        self.tracker_path = filepath

        return filepath

    def add_sample_data(self, filepath):
        """Add sample data to demonstrate the tracker"""
        wb = load_workbook(filepath)
        ws = wb['Project Status']

        sample_tasks = [
            ['T001', 'Chandler Lewis', 'SpacePlan JV legal documentation', 'In Progress', 'Critical', 75, '', '', '2025-11-10', 'Gmail: Thread with SpacePlan legal team', 'Finalize term sheet by Nov 18', '2025-11-05', '2025-11-14', 'No'],
            ['T002', 'Team Member', 'GenIP integration testing', 'Blocked', 'High', 40, 'Waiting on API credentials from GenIP', '', '2025-11-08', 'Drive: Meeting notes from Nov 8', 'Follow up with GenIP tech team', '2025-11-01', '2025-11-14', 'No'],
            ['T003', 'Team Member', 'CNEN portfolio assessment report', 'In Progress', 'High', 60, '', '', '2025-11-12', 'Calendar: CNEN sync meeting', 'Complete technology matrix by Nov 20', '2025-11-02', '2025-11-14', 'No'],
            ['T004', 'Chandler Lewis', 'Board meeting presentation prep', 'Not Started', 'Critical', 0, '', '', '2025-11-14', 'Calendar: Board meeting Nov 17', 'Compile partnership updates', '2025-11-13', '2025-11-14', 'No'],
            ['T005', 'Team Member', 'NanoBioPlus partnership documentation', 'Done', 'Medium', 100, '', '', '2025-11-13', 'Gmail: Signed agreement received', 'Archive completed docs', '2025-10-15', '2025-11-14', 'No'],
        ]

        for row_idx, task_data in enumerate(sample_tasks, start=2):
            for col_idx, value in enumerate(task_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Add alignment for better readability
                if col_idx in [3, 7, 10, 11]:  # Text columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                elif col_idx == 6:  # Progress %
                    cell.number_format = '0'
                elif col_idx in [9, 12, 13]:  # Date columns
                    if value and value != '':
                        cell.value = datetime.strptime(value, '%Y-%m-%d')
                        cell.number_format = 'YYYY-MM-DD'

        # Add cell comments with source context
        from openpyxl.comments import Comment

        comments = {
            'J2': 'Source: Gmail thread with SpacePlan legal\nDate: Nov 10, 2025\nSubject: JV Documentation Review\nKey point: Term sheet markup completed',
            'J3': 'Source: Google Drive meeting notes\nFile: GenIP Integration Sync - Nov 8\nStatus: Blocked on API access\nContact: tech@genip.com',
            'J4': 'Source: Calendar meeting\nMeeting: CNEN Portfolio Review\nDate: Nov 12, 2025\nAttendees: Chandler, CNEN team\nNext: Technology assessment matrix',
            'J5': 'Source: Calendar event\nEvent: 360 Annual Board Meeting\nDate: Nov 17, 2025\nLocation: Seattle office\nPrep: Partnership updates needed',
        }

        for cell_ref, comment_text in comments.items():
            cell = ws[cell_ref]
            cell.comment = Comment(comment_text, 'System')

        wb.save(filepath)

        return filepath

    def collect_data_from_sources(self, days_back=7):
        """
        Collect status information from available sources.
        Returns list of task dictionaries.
        """
        tasks = []
        cutoff_date = datetime.now() - timedelta(days=days_back)

        # This will be filled in with actual data collection logic
        # For now, returning structure that shows what we'll collect

        collection_plan = {
            'gmail': {
                'queries': [
                    'status update',
                    'blocker',
                    'blocked on',
                    'update on',
                    'working on',
                    'completed',
                    'finished',
                ],
                'date_filter': cutoff_date.strftime('%Y/%m/%d'),
                'extract': ['sender', 'subject', 'date', 'key_phrases']
            },
            'calendar': {
                'lookback_days': days_back,
                'extract': ['attendees', 'title', 'notes', 'description', 'action_items']
            },
            'drive': {
                'folder': 'A - Processed Meeting Deliverables',
                'date_filter': cutoff_date,
                'extract': ['title', 'content', 'last_modified', 'owner']
            }
        }

        return tasks, collection_plan

def main():
    """Main execution function"""
    print("360 Project Status Tracker Builder")
    print("=" * 50)

    builder = ProjectTrackerBuilder()

    # Create the tracker template
    print("\n1. Creating tracker template...")
    tracker_path = builder.create_tracker_template()
    print(f"   ✓ Template created at: {tracker_path}")

    # Add sample data
    print("\n2. Adding sample data...")
    builder.add_sample_data(tracker_path)
    print(f"   ✓ Sample data added")

    # Show collection plan
    print("\n3. Data collection plan:")
    tasks, plan = builder.collect_data_from_sources()
    print(f"   • Gmail: {len(plan['gmail']['queries'])} search queries")
    print(f"   • Calendar: Last {plan['calendar']['lookback_days']} days")
    print(f"   • Drive: Folder '{plan['drive']['folder']}'")

    print("\n" + "=" * 50)
    print(f"Tracker ready: {tracker_path}")
    print("\nNext steps:")
    print("  1. Review sample data in tracker")
    print("  2. Connect live data sources")
    print("  3. Configure update mode (on-demand or recurring)")

if __name__ == "__main__":
    main()
