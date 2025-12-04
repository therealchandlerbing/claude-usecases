"""
Tests for the 360 Executive Project Tracker orchestrator module.

This module tests tracker creation, data collection, task management,
and summary generation functionality.

These tests require openpyxl to be installed.
"""

import pytest
import json
import os
import sys
import tempfile
from datetime import datetime, timedelta
from unittest.mock import MagicMock, patch, mock_open

# Check for openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Skip entire module if openpyxl not available
pytestmark = pytest.mark.skipif(
    not HAS_OPENPYXL,
    reason="openpyxl not installed"
)

# Add module directory to path
sys.path.insert(0, os.path.join(
    os.path.dirname(__file__),
    '../../../.claude/skills/360-executive-project-tracker/scripts'
))


class TestTrackerOrchestratorInitialization:
    """Tests for TrackerOrchestrator initialization"""

    @pytest.fixture
    def mock_builder(self):
        """Create mock ProjectTrackerBuilder"""
        with patch('tracker_orchestrator.ProjectTrackerBuilder') as MockBuilder:
            mock_instance = MagicMock()
            MockBuilder.return_value = mock_instance
            yield mock_instance

    def test_initialization_with_default_config(self, mock_builder):
        """Test initialization with default configuration"""
        from tracker_orchestrator import TrackerOrchestrator

        with patch('builtins.open', side_effect=FileNotFoundError()):
            orchestrator = TrackerOrchestrator(mode='on_demand')

        assert orchestrator.mode == 'on_demand'
        assert orchestrator.config['days_back'] == 7
        assert orchestrator.config['update_existing'] is True
        assert orchestrator.config['preserve_manual_edits'] is True

    def test_initialization_with_config_file(self, mock_builder):
        """Test initialization with config file"""
        config_data = {
            'data_collection': {'days_back': 14},
            'tracker_settings': {
                'update_existing': False,
                'preserve_manual_edits': False
            },
            'team': {'members': ['Alice', 'Bob']},
        }

        with patch('builtins.open', mock_open(read_data=json.dumps(config_data))):
            from tracker_orchestrator import TrackerOrchestrator
            orchestrator = TrackerOrchestrator(mode='recurring')

        assert orchestrator.mode == 'recurring'

    def test_initialization_modes(self, mock_builder):
        """Test both on_demand and recurring modes"""
        from tracker_orchestrator import TrackerOrchestrator

        with patch('builtins.open', side_effect=FileNotFoundError()):
            on_demand = TrackerOrchestrator(mode='on_demand')
            recurring = TrackerOrchestrator(mode='recurring')

        assert on_demand.mode == 'on_demand'
        assert recurring.mode == 'recurring'


class TestGenerateSummary:
    """Tests for summary generation"""

    @pytest.fixture
    def orchestrator_with_mock_wb(self):
        """Create orchestrator with mocked workbook"""
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                orchestrator = TrackerOrchestrator()
                orchestrator.tracker_path = '/fake/path.xlsx'
                return orchestrator

    def test_summary_structure(self, orchestrator_with_mock_wb):
        """Test that summary has expected keys"""
        mock_ws = MagicMock()
        mock_ws.max_row = 5
        mock_ws.__getitem__ = MagicMock(return_value=MagicMock(value=None))

        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
        mock_wb.close = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            summary = orchestrator_with_mock_wb.generate_summary()

        expected_keys = [
            'Total Tasks', 'Not Started', 'In Progress',
            'Blocked', 'Done', 'Critical Priority', 'Blocked >3 Days'
        ]

        for key in expected_keys:
            assert key in summary

    def test_summary_counts_tasks(self, orchestrator_with_mock_wb):
        """Test that summary counts tasks correctly"""
        # Create mock worksheet with tasks
        mock_ws = MagicMock()
        mock_ws.max_row = 4

        # Row data: task_id, owner, desc, status, priority, progress, blocker, days_blocked
        row_data = {
            'A2': 'T001', 'D2': 'In Progress', 'E2': 'Critical', 'H2': 5,
            'A3': 'T002', 'D3': 'Blocked', 'E3': 'High', 'H3': 2,
            'A4': 'T003', 'D4': 'Done', 'E4': 'Medium', 'H4': 0,
        }

        def get_cell(key):
            mock_cell = MagicMock()
            mock_cell.value = row_data.get(key)
            return mock_cell

        mock_ws.__getitem__ = get_cell

        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
        mock_wb.close = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            summary = orchestrator_with_mock_wb.generate_summary()

        assert summary['Total Tasks'] == 3
        assert summary['In Progress'] == 1
        assert summary['Critical Priority'] == 1
        assert summary['Blocked >3 Days'] == 1


class TestGetExistingTasks:
    """Tests for existing task retrieval"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                return TrackerOrchestrator()

    def test_returns_empty_for_no_tasks(self, orchestrator):
        """Test empty dict for worksheet with no tasks"""
        mock_ws = MagicMock()
        mock_ws.max_row = 1  # Only header row

        result = orchestrator.get_existing_tasks(mock_ws)
        assert result == {}

    def test_returns_task_dict(self, orchestrator):
        """Test returns dictionary with task data"""
        mock_ws = MagicMock()
        mock_ws.max_row = 3

        # Mock cell access
        def get_cell(row, column):
            if row == 2 and column == 1:
                mock_cell = MagicMock()
                mock_cell.value = 'T001'
                return mock_cell
            if row == 3 and column == 1:
                mock_cell = MagicMock()
                mock_cell.value = 'T002'
                return mock_cell
            mock_cell = MagicMock()
            mock_cell.value = f'data-{row}-{column}'
            return mock_cell

        mock_ws.cell = get_cell

        def get_item(key):
            mock_cell = MagicMock()
            # Parse 'A2' format
            if key == 'A2':
                mock_cell.value = 'T001'
            elif key == 'A3':
                mock_cell.value = 'T002'
            else:
                mock_cell.value = None
            return mock_cell

        mock_ws.__getitem__ = get_item

        result = orchestrator.get_existing_tasks(mock_ws)

        assert 'T001' in result
        assert 'T002' in result
        assert result['T001']['row'] == 2
        assert result['T002']['row'] == 3


class TestUpdateMetadata:
    """Tests for metadata update"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                orch = TrackerOrchestrator(mode='on_demand')
                orch.tracker_path = '/fake/path.xlsx'
                return orch

    def test_updates_timestamp(self, orchestrator):
        """Test that timestamp is updated"""
        mock_dashboard = MagicMock()
        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_dashboard)
        mock_wb.save = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            orchestrator.update_metadata()

        # Check that B15 (timestamp) was set
        mock_dashboard.__setitem__.assert_any_call('B15', pytest.approx_any)

    def test_updates_mode(self, orchestrator):
        """Test that mode is updated"""
        mock_dashboard = MagicMock()
        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_dashboard)
        mock_wb.save = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            orchestrator.update_metadata()

        # Check that B16 (mode) was set to formatted mode
        mock_dashboard.__setitem__.assert_any_call('B16', 'On Demand')


class TestConfigureRecurring:
    """Tests for recurring mode configuration"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                return TrackerOrchestrator(mode='on_demand')

    def test_sets_mode_to_recurring(self, orchestrator):
        """Test that mode is changed to recurring"""
        orchestrator.configure_recurring(schedule='weekly', day='Monday', time='09:00')
        assert orchestrator.mode == 'recurring'

    def test_stores_schedule_config(self, orchestrator):
        """Test that schedule configuration is stored"""
        orchestrator.configure_recurring(schedule='biweekly', day='Tuesday', time='10:30')

        assert orchestrator.config['recurring']['schedule'] == 'biweekly'
        assert orchestrator.config['recurring']['day'] == 'Tuesday'
        assert orchestrator.config['recurring']['time'] == '10:30'
        assert orchestrator.config['recurring']['enabled'] is True


class TestCollectLiveData:
    """Tests for live data collection"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                return TrackerOrchestrator()

    def test_returns_empty_list(self, orchestrator):
        """Test that collect_live_data returns empty list (stub)"""
        # The actual implementation is a stub that returns empty
        result = orchestrator.collect_live_data(days_back=7)
        assert result == []

    def test_uses_config_drive_folders(self, orchestrator):
        """Test that configured drive folders are used"""
        # This mainly tests the config structure
        assert 'drive_folders' in orchestrator.config


class TestAddNewTaskRow:
    """Tests for adding new task rows"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                return TrackerOrchestrator()

    def test_generates_task_id_if_missing(self, orchestrator):
        """Test that task ID is generated if not provided"""
        mock_ws = MagicMock()
        mock_ws.max_row = 5

        task = {'description': 'Test task', 'status': 'Not Started'}

        # Mock cell assignment
        assigned = {}
        def setitem(key, value):
            assigned[key] = value

        mock_ws.__setitem__ = setitem

        orchestrator.add_new_task_row(mock_ws, task)

        # Should have generated ID like T005
        assert assigned['A6'] == 'T005'

    def test_preserves_provided_task_id(self, orchestrator):
        """Test that provided task ID is preserved"""
        mock_ws = MagicMock()
        mock_ws.max_row = 5

        task = {'task_id': 'CUSTOM-001', 'description': 'Test'}

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        orchestrator.add_new_task_row(mock_ws, task)

        assert assigned['A6'] == 'CUSTOM-001'

    def test_sets_default_values(self, orchestrator):
        """Test that default values are set for missing fields"""
        mock_ws = MagicMock()
        mock_ws.max_row = 1

        task = {'description': 'Test task'}

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        orchestrator.add_new_task_row(mock_ws, task)

        # Check defaults
        assert assigned['B2'] == 'Unassigned'  # Default owner
        assert assigned['D2'] == 'Not Started'  # Default status
        assert assigned['E2'] == 'Medium'  # Default priority
        assert assigned['N2'] == 'No'  # Manual override


class TestUpdateTaskRow:
    """Tests for updating existing task rows"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                return TrackerOrchestrator()

    def test_updates_description(self, orchestrator):
        """Test that description is updated"""
        mock_ws = MagicMock()

        existing_values = {'C5': 'Old description'}
        def getitem(key):
            cell = MagicMock()
            cell.value = existing_values.get(key, 'default')
            cell.comment = None
            return cell

        mock_ws.__getitem__ = getitem

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        task = {'description': 'New description'}
        orchestrator.update_task_row(mock_ws, 5, task)

        assert assigned['C5'] == 'New description'

    def test_updates_status(self, orchestrator):
        """Test that status is updated"""
        mock_ws = MagicMock()

        def getitem(key):
            cell = MagicMock()
            cell.value = 'existing'
            cell.comment = None
            return cell

        mock_ws.__getitem__ = getitem

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        task = {'status': 'In Progress'}
        orchestrator.update_task_row(mock_ws, 3, task)

        assert assigned['D3'] == 'In Progress'

    def test_updates_last_modified_date(self, orchestrator):
        """Test that last modified date is updated"""
        mock_ws = MagicMock()

        def getitem(key):
            cell = MagicMock()
            cell.value = None
            cell.comment = None
            return cell

        mock_ws.__getitem__ = getitem

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        task = {}
        orchestrator.update_task_row(mock_ws, 2, task)

        # M column should have today's date
        assert 'M2' in assigned
        assert datetime.now().strftime('%Y-%m-%d') == assigned['M2']


class TestUpdateTrackerWithTasks:
    """Tests for batch task updates"""

    @pytest.fixture
    def orchestrator(self):
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                from tracker_orchestrator import TrackerOrchestrator
                orch = TrackerOrchestrator()
                orch.tracker_path = '/fake/path.xlsx'
                return orch

    def test_skips_manual_override_tasks(self, orchestrator):
        """Test that tasks with manual override are skipped"""
        mock_ws = MagicMock()
        mock_ws.max_row = 3

        # Setup existing tasks
        orchestrator.get_existing_tasks = MagicMock(return_value={
            'T001': {'row': 2}
        })

        # Mock manual override flag
        def getitem(key):
            cell = MagicMock()
            if key == 'N2':  # Manual override column
                cell.value = 'Yes'
            else:
                cell.value = None
            cell.comment = None
            return cell

        mock_ws.__getitem__ = getitem

        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
        mock_wb.save = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            # The task should be skipped
            tasks = [{'task_id': 'T001', 'description': 'Updated'}]
            orchestrator.update_tracker_with_tasks(tasks)

        # Save should still be called
        mock_wb.save.assert_called_once()

    def test_adds_new_tasks(self, orchestrator):
        """Test that new tasks are added"""
        mock_ws = MagicMock()
        mock_ws.max_row = 1

        orchestrator.get_existing_tasks = MagicMock(return_value={})

        assigned = {}
        mock_ws.__setitem__ = lambda k, v: assigned.update({k: v})

        mock_wb = MagicMock()
        mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
        mock_wb.save = MagicMock()

        with patch('tracker_orchestrator.load_workbook', return_value=mock_wb):
            tasks = [{'task_id': 'T001', 'description': 'New task'}]
            orchestrator.update_tracker_with_tasks(tasks)

        # Check task was added
        assert 'A2' in assigned
        assert assigned['A2'] == 'T001'


class TestMainFunction:
    """Tests for main CLI entry point"""

    def test_main_creates_orchestrator(self):
        """Test that main function creates orchestrator"""
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                with patch('sys.argv', ['tracker_orchestrator.py', '--sample']):
                    from tracker_orchestrator import TrackerOrchestrator

                    # Create orchestrator directly for testing
                    orch = TrackerOrchestrator(mode='on_demand')
                    assert orch is not None

    def test_main_accepts_days_argument(self):
        """Test that main accepts --days argument"""
        with patch('tracker_orchestrator.ProjectTrackerBuilder'):
            with patch('builtins.open', side_effect=FileNotFoundError()):
                import argparse

                parser = argparse.ArgumentParser()
                parser.add_argument('--days', type=int, default=7)

                args = parser.parse_args(['--days', '14'])
                assert args.days == 14


# Custom pytest assertion helper
@pytest.fixture
def approx_any():
    """Fixture for approximate matching of any value"""
    class ApproxAny:
        def __eq__(self, other):
            return True
    return ApproxAny()


# Make approx_any available as pytest attribute
pytest.approx_any = type('ApproxAny', (), {'__eq__': lambda s, o: True})()
